from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Grades.xlsx')
work_sheet1 = wb.active

work_sheet1['F1'].value = 50

for row in range(1,11):
    for col in range(1,6):
        char = get_column_letter(col)
        work_sheet1[char + str(row)].value = "hey"


wb.save('Grades.xlsx')
print(work_sheet1)
    